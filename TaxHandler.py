from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from datetime import datetime
import TaxHandlerComponents as thc
import json
import os

# Read JSON from a file
with open('config.json', 'r', encoding='utf-8') as file:
    data = json.load(file)  # Use json.load() to read from the file

# Access the wb_name key in the JSON content
wb_name = data['config']['wb_name']

wb = load_workbook(wb_name)

import sys
from PyQt5.QtWidgets import (QApplication, QWidget, QVBoxLayout, 
                             QPushButton, QLabel, QLineEdit, 
                             QStackedWidget, QFormLayout, 
                             QScrollArea, QGroupBox, QHBoxLayout)
import win32com.client as win32
from datetime import datetime

class ExcelPrinter(QWidget):
    def __init__(self, branches):
        super().__init__()
        self.branches = branches
        self.selected_branch = None
        self.date = []
        self.amount = None
        self.price = None
        self.totalBeforeDiscount = None
        self.discount = None
        self.netTotal = None
        self.vat = None
        self.total = None
        self.sales_id = None
        self.initUI()

    def initUI(self):
        self.layout = QVBoxLayout()
        self.stacked_widget = QStackedWidget()

        # Page 1: Branch Selection
        self.page1 = QWidget()
        self.page1_layout = QVBoxLayout()
        self.label = QLabel("Choose a branch for printing the receipt", self.page1)

        # Create a scroll area for branch buttons
        self.scroll_area = QScrollArea(self.page1)
        self.scroll_area.setWidgetResizable(True)

        # Create a container widget for the scroll area
        self.scroll_container = QGroupBox(self.scroll_area)
        self.scroll_layout = QVBoxLayout(self.scroll_container)

        # Add buttons for each branch in a scrollable layout
        for branch in self.branches:
            button = QPushButton(branch, self.scroll_container)
            button.clicked.connect(lambda checked, b=branch: self.select_branch(b))
            self.scroll_layout.addWidget(button)

        self.scroll_area.setWidget(self.scroll_container)
        self.page1_layout.addWidget(self.label)
        self.page1_layout.addWidget(self.scroll_area)

        self.page1.setLayout(self.page1_layout)

        # Page 2: Input Details
        self.page2 = QWidget()
        self.page2_layout = QFormLayout()
        self.date_input = QLineEdit(self.page2)
        self.month_input = QLineEdit(self.page2)
        self.year_input = QLineEdit(self.page2)
        self.amount_input = QLineEdit(self.page2)
        self.price_input = QLineEdit(self.page2)
        self.submit_button = QPushButton("Submit", self.page2)
        self.submit_button.clicked.connect(self.submit_details)

        self.page2_layout.addRow("Enter Date:", self.date_input)
        self.page2_layout.addRow("Enter Month:", self.month_input)
        self.page2_layout.addRow("Enter Year:", self.year_input)
        self.page2_layout.addRow("Enter Amount:", self.amount_input)
        self.page2_layout.addRow("Enter Price:", self.price_input)
        self.page2_layout.addWidget(self.submit_button)

        self.page2.setLayout(self.page2_layout)

        # Page 3: Printing Confirmation
        self.page3 = QWidget()
        self.page3_layout = QVBoxLayout()
        self.print_label = QLabel("Printing the Excel sheet...", self.page3)
        self.page3_layout.addWidget(self.print_label)
        self.page3.setLayout(self.page3_layout)

        # Add pages to stacked widget
        self.stacked_widget.addWidget(self.page1)
        self.stacked_widget.addWidget(self.page2)
        self.stacked_widget.addWidget(self.page3)

        self.layout.addWidget(self.stacked_widget)
        self.setLayout(self.layout)
        self.setWindowTitle("Receipt Printer")
        self.setGeometry(300, 300, 400, 300)

    def select_branch(self, branch):
        self.selected_branch = branch
        self.label.setText(f"Selected Branch: {branch}")
        self.stacked_widget.setCurrentIndex(1)  # Go to page 2

    def submit_details(self):
        self.date.append(self.date_input.text())
        self.date.append(self.month_input.text())
        self.date.append(self.year_input.text())
        self.amount = int(self.amount_input.text())
        self.price = float(self.price_input.text())
        self.calculate_totals()

        # Proceed to the printing page
        self.stacked_widget.setCurrentIndex(2)
        self.print_excel_file("C:\\path\\to\\your\\file.xlsx")  # Replace with actual path

    def calculate_totals(self):
        totalBeforeDiscount = self.amount * self.price
        discount = 0.05 * totalBeforeDiscount
        netTotal = totalBeforeDiscount - discount
        vat = netTotal * 0.07
        total = netTotal + vat
        
        # Store calculated values
        self.totalBeforeDiscount = totalBeforeDiscount
        self.discount = discount
        self.netTotal = netTotal
        self.vat = vat
        self.total = total

    def print_excel_file(self, file_path):
        try:
            excel = win32.Dispatch("Excel.Application")
            excel.Visible = False
            workbook = excel.Workbooks.Open(file_path)

            # Select the first sheet for demonstration
            sheet = workbook.Sheets(1)

            # Set your printer here (adjust based on your printer setup)
            excel.ActivePrinter = "Microsoft Print to PDF"

            # Here you would replace this with the actual logic to get sales_id
            self.sales_id = 1  # Dummy value; implement actual ID logic

            indiv_row = [self.sales_id, datetime(int(self.date[2]), int(self.date[1]), int(self.date[0])),
                         self.amount, self.price, self.totalBeforeDiscount,
                         self.discount, self.netTotal, self.vat, self.total]

            # Logic for writing to Excel goes here (like your provided code)
            # Append the data to row i by writing in cells
            latest_row_num = self.find_last_row_with_data(sheet)  # Implement this function

            for col_num, value in enumerate(indiv_row, start=1):
                new_cell = sheet.cell(row=latest_row_num + 1, column=col_num, value=value)
                # Copy individual styles from the old cell to the new cell (as per your example)
                # Implement cell style copying as needed

            workbook.Save()  # Save the workbook
            workbook.Close(SaveChanges=True)
            excel.Quit()

            print("Printed successfully.")
        except Exception as e:
            print(f"An error occurred: {e}")

    def find_last_row_with_data(self, sheet):
        # Implement your logic to find the last row with data
        pass  # Replace with actual logic

if __name__ == '__main__':
    branches = data['config']['address'].keys()

    app = QApplication(sys.argv)
    ex = ExcelPrinter(branches)
    ex.show()
    sys.exit(app.exec_())


# while True:





#     # Modifying Sales sheet

#     sheet = thc.ask_for_sheet(wb)

#     if not sheet:
#         continue

#     latest_row_num = thc.find_last_row_with_data(sheet)

#     if latest_row_num:
#         latest_row_indiv = [cell.value for cell in sheet[latest_row_num]]
#         print(f"Latest row with data is row number: {latest_row_num}")
#         print(f"Data in the latest row: {latest_row_indiv}")
#     else:
#         print("No data found in the sheet.")

#     date = []
#     date.append(input("Enter Date: \n"))
#     date.append(input("Enter Month: \n"))
#     date.append(input("Enter Year: \n"))

#     amount = int(input("Enter Amount: \n"))

#     price = float(input("Enter Price: \n"))
#     totalBeforeDiscount = amount * price

#     discount = 0.05 * totalBeforeDiscount
#     netTotal = totalBeforeDiscount - discount
#     vat = netTotal * 0.07
#     total = netTotal + vat

#     sales_id = thc.iterate_id(latest_row_indiv[0])

#     indiv_row = [sales_id, datetime(int(date[2]), int(date[1]), int(date[0])), amount, price, totalBeforeDiscount, discount, netTotal, vat, total]

#     # Append the data to row i by writing in cells
#     for col_num, value in enumerate(indiv_row, start=1):
#         # Get the cell in the new row
#         new_cell = sheet.cell(row=latest_row_num + 1, column=col_num, value=value)

#         # Get the corresponding cell in the previous row
#         old_cell = sheet.cell(row=latest_row_num, column=col_num)

#         # Copy individual styles from the old cell to the new cell
#         new_cell.font = Font(name=old_cell.font.name, size=old_cell.font.size, bold=old_cell.font.bold, italic=old_cell.font.italic)
#         new_cell.border = Border(left=old_cell.border.left, right=old_cell.border.right, top=old_cell.border.top, bottom=old_cell.border.bottom)
#         new_cell.number_format = old_cell.number_format
#         new_cell.alignment = Alignment(horizontal=old_cell.alignment.horizontal, vertical=old_cell.alignment.vertical)


#     # wb.save(wb_name)
#     print(f"Data successfully inserted into {sheet.title} sheet with ID: {indiv_row[0]}")




#     # Modifying Tax sheet

#     sheet = data['config']['tax_sheet_name']

#     if sheet not in wb.sheetnames:
#         print(f"Sheet '{sheet}' not found in the workbook.")
#     else:
#         print("Sheet selected: ", sheet)
#         sheet = wb[sheet]

#     latest_row_num_tax = thc.find_last_row_with_data(sheet)

#     if latest_row_num_tax:
#         latest_row_tax = [cell.value for cell in sheet[latest_row_num_tax]]
#         print(f"Latest row with data is row number: {latest_row_num_tax}")
#         print(f"Data in the latest row: {latest_row_tax}")
#     else:
#         print("No data found in the sheet.")

#     tax_num = int(latest_row_tax[1]) + 1

#     tax_row = [latest_row_tax[0] + 1, tax_num, datetime(int(date[2]), int(date[1]), int(date[0])), sales_id, totalBeforeDiscount, discount, netTotal, vat, total]

#     # Append the data to row i by writing in cells
#     for col_num, value in enumerate(tax_row, start=1):
#         # Get the cell in the new row
#         new_cell = sheet.cell(row=latest_row_num_tax + 1, column=col_num, value=value)

#         # Get the corresponding cell in the previous row
#         old_cell = sheet.cell(row=latest_row_num_tax, column=col_num)

#         # Copy individual styles from the old cell to the new cell
#         new_cell.font = Font(name=old_cell.font.name, size=old_cell.font.size, bold=old_cell.font.bold, italic=old_cell.font.italic)
#         new_cell.border = Border(left=old_cell.border.left, right=old_cell.border.right, top=old_cell.border.top, bottom=old_cell.border.bottom)
#         new_cell.number_format = old_cell.number_format
#         new_cell.alignment = Alignment(horizontal=old_cell.alignment.horizontal, vertical=old_cell.alignment.vertical)

#     # wb.save(wb_name)
#     print(f"Data successfully inserted into {sheet} sheet with ID: {tax_row[0]}")




#     # Modifying Vat sheet

#     sheet = data['config']['vat_sheet_name']

#     if sheet not in wb.sheetnames:
#         print(f"Sheet '{sheet}' not found in the workbook.")
#     else:
#         print("Sheet selected: ", sheet)
#         sheet = wb[sheet]

#     # real / copy
#     sheet.cell(row=2, column=6, value='(ต้นฉบับ)')

#     # tax number
#     sheet.cell(row=3, column=7, value=tax_num)
#     sheet.cell(row=3, column=16, value=tax_num)
    
#     # sales number
#     sheet.cell(row=4, column=7, value=sales_id)
#     sheet.cell(row=4, column=16, value=sales_id)

#     # print date
#     sheet.cell(row=5, column=7, value=thc.return_date_now())
#     sheet.cell(row=5, column=16, value=thc.return_date_now())

#     # amount
#     sheet.cell(row=7, column=5, value=amount)
#     sheet.cell(row=7, column=14, value=amount)

#     # price
#     sheet.cell(row=7, column=6, value=price)
#     sheet.cell(row=7, column=15, value=price)

#     # total before discount
#     sheet.cell(row=7, column=7, value=totalBeforeDiscount)
#     sheet.cell(row=7, column=16, value=totalBeforeDiscount)
#     sheet.cell(row=9, column=7, value=totalBeforeDiscount)
#     sheet.cell(row=9, column=16, value=totalBeforeDiscount)

#     # discount
#     sheet.cell(row=10, column=7, value=discount)
#     sheet.cell(row=10, column=16, value=discount)

#     # net total
#     sheet.cell(row=11, column=7, value=netTotal)
#     sheet.cell(row=11, column=16, value=netTotal)

#     # vat
#     sheet.cell(row=12, column=7, value=vat)
#     sheet.cell(row=12, column=16, value=vat)

#     # total
#     sheet.cell(row=13, column=7, value=total)
#     sheet.cell(row=13, column=16, value=total)
    
#     # address
#     sheet.cell(row=4, column=2, value=thc.return_address(sales_id))
#     sheet.cell(row=4, column=11, value=thc.return_address(sales_id))

#     wb.save(wb_name)
#     print(f"Data successfully saved in {wb_name}.")

#     # file_path = os.path.join(os.getcwd(), wb_name) 
#     file_path = os.path.abspath(wb_name)

#     print(file_path)
#     file_path = file_path.encode('utf-8').decode('utf-8')

#     thc.print_excel_sheet(file_path, sheet.title)

#     # real / copy
#     sheet.cell(row=2, column=6, value='(สำเนา)')

#     thc.print_excel_sheet(file_path, sheet.title)
#     print("Printing the Excel sheet...")